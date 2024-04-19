import csv
import openpyxl
import time
import datetime
import warnings
import mysql.connector

def get_end_time():
    now = datetime.datetime.now()
    return datetime.datetime(now.year, now.month, now.day, 16, 41, 0)

def countdown(stop):
    fileNameCustomerClaim = "file"

    while True:
        difference = stop - datetime.datetime.now()
        count_hours, rem = divmod(difference.seconds, 3600)
        count_minutes, count_seconds = divmod(rem, 60)
        if difference.days < 0:
            stop += datetime.timedelta(days=1)
        elif difference.days == 0 and count_hours == 0 and count_minutes == 0 and count_seconds == 0:
            print("Countdown finished!")

            # For SQL
            try:
                # Establish a connection to MySQL database
                conn = mysql.connector.connect(
                    host='localhost',
                    database='qa_claimdb_conversion',
                    user='sys_user',
                    password='P@ssword'
                )
                cursor = conn.cursor()

                # Execute the SQL First Insert Statent
                customerClaimsql = "INSERT INTO tblconversion_logs (CONVERSION_FILE_NAME,CONVERSION_DATETIME) VALUES (%s, %s)"
                customerClaimval = (fileNameCustomerClaim, stop)
                cursor.execute(customerClaimsql, customerClaimval)

                conn.commit()

                print(cursor.rowcount, "record inserted.")
            except mysql.connector.Error as error:
                print("Failed to insert record into MySQL table: {}".format(error))
            finally:
                # Close the connection to MySQL
                if conn.is_connected():
                    cursor.close()
                    conn.close()
                    print("MySQL connection is closed")
            # For SQL

            print("Initializing Script...")
            time.sleep(1)
            convert_to_csvCustomerClaimDB()
         
            print("The excel is now converted. The cycle will continue, thank you!")
            stop += datetime.timedelta(days=1)
        print('Countdown: '
              + str(difference.days) + " day(s) "
              + str(count_hours) + " hour(s) "
              + str(count_minutes) + " minute(s) "
              + str(count_seconds) + " second(s) "
              )
        time.sleep(1)
        stop = stop

def convert_to_csvCustomerClaimDB():
    warnings.simplefilter(action='ignore', category=UserWarning)
    # Load the Excel workbook
    wb = openpyxl.load_workbook(r'\\network_folder\file.xlsx')
    # Specify the name of the sheet you want to select
    sheet_name = 'CUSTOMER CLAIM'
    sheet = wb[sheet_name]
    # Get the index of the "PART NAME" column
    part_name_col = None
    header = []
    for i, row in enumerate(sheet.iter_rows()):
        if i == 1:
            header = [cell.value for cell in row]
        for j, cell in enumerate(row):
            if cell.value == "PART NAME":
                part_name_col = j
                break
        if part_name_col is not None:
            break
    # Open a new CSV file for writing
    with open(r'\\network_folder\file.csv', 'w', newline='') as f:
        csv_writer = csv.writer(f)
        # Write the header to the CSV file
        csv_writer.writerow(header)
        # Write each row of the sheet to the CSV file
        for row in sheet.iter_rows():
            if row[part_name_col].value == "Meter Assy Combination" or row[part_name_col].value == "Panel Assy, Center Integration":
                csv_writer.writerow([cell.value for cell in row])
    print("Success")



print("The Script is Running")
countdown(get_end_time())


